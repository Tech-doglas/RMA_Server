# app/routes/main_routes.py
from flask import Blueprint, request, jsonify, make_response, send_file
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from io import BytesIO
from collections import defaultdict
from app.models import get_db_connection

from app.routes.laptop.laptop_item_routes import laptop_item_bp
from app.routes.laptop.laptop_sales_routes import laptop_sales_bp

laptop_bp = Blueprint('laptop', __name__)
laptop_bp.register_blueprint(laptop_item_bp, url_prefix='/item')
laptop_bp.register_blueprint(laptop_sales_bp, url_prefix='/sales')

condition_mapping = {
    'Back to New': 'N',
    'Grade A': 'A',
    'Grade B': 'B',
    'Grade C': 'C',
    'Grade F': 'F'
}

@laptop_bp.route('/api/laptops/search', methods=['POST'])
def api_search_laptops():
    try:
        filters = request.json

        conn = get_db_connection()
        cursor = conn.cursor()

        query = "SELECT * FROM RMA_laptop_sheet WHERE 1=1"
        params = []

        # === Serial Number ===
        serial_number = filters.get('serialNumber', '').strip()
        if serial_number:
            query += " AND SerialNumber LIKE ?"
            params.append(f"%{serial_number}%")

        # === Model ===
        model = filters.get('model', '').strip()
        if model:
            query += " AND Model LIKE ?"
            params.append(f"%{model}%")

        # === Brand ===
        brand = filters.get('brand', '')
        if isinstance(brand, list):
            brand = brand[0] if brand else ''
        brand = brand.strip() if brand else ''
        if brand:
            query += " AND Brand = ?"
            params.append(brand)

        condition_list = filters.get('conditions', [])
        db_conditions = [condition_mapping.get(c, c) for c in condition_list]
        if db_conditions:
            placeholders = ','.join(['?'] * len(db_conditions))
            query += f" AND Condition IN ({placeholders})"
            params.extend(db_conditions)

        # === Stock ===
        stock_list = filters.get('stock', [])
        stock_conditions = []

        if 'SOLD' in stock_list:
            stock_conditions.append("Stock = ?")
            params.append('SOLD')

        if 'In Stock' in stock_list:
            stock_conditions.append("(Stock IS NULL OR Stock = '')")  # no param

        if stock_conditions:
            query += " AND (" + " OR ".join(stock_conditions) + ")"

        # === Tech Done ===
        tech_done_list = filters.get('techDone', [])
        tech_conditions = []

        if 'Done' in tech_done_list:
            tech_conditions.append("TechDone = ?")
            params.append(1)

        if 'Not yet' in tech_done_list:
            tech_conditions.append("TechDone = ?")
            params.append(0)
            tech_conditions.append("TechDone IS NULL")  # no param

        if tech_conditions:
            query += " AND (" + " OR ".join(tech_conditions) + ")"

        # === Execute ===
        cursor.execute(query, params)
        data = cursor.fetchall()

        result = []
        if data:
            columns = [col[0] for col in cursor.description]
            result = [dict(zip(columns, row)) for row in data]

        conn.close()
        return jsonify(result)

    except Exception as e:
        import traceback
        traceback.print_exc()
        print("❌ ERROR:", str(e))
        return jsonify({'error': str(e)}), 500



@laptop_bp.route('/api/users', methods=['GET'])
def get_users():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM RMA_user")
        data = cursor.fetchall()
        conn.close()

        results = [row[0] for row in data]
        return jsonify(results)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@laptop_bp.route('/laptop_SRP', methods=['GET'])
def laptop_SRP():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT SaleDate, Stock, Condition FROM RMA_laptop_sheet")
        rows = cursor.fetchall()
        conn.close()

        valid_grades = {'N', 'A', 'B', 'C', 'F'}
        grades_seen = set()

        # Track sales by month and grade
        sold_by_month_and_grade = defaultdict(lambda: defaultdict(int))

        # Count total inventory by grade
        total_inventory_by_grade = defaultdict(int)
        
        # Track all months for ordering
        all_months = set()

        # First pass: Count total inventory and identify all grades
        for sale_date, stock, grade in rows:
            grade = (grade or '').strip().upper()
            if grade not in valid_grades:
                grade = 'Unknown'
            grades_seen.add(grade)
            
            # Count all items in inventory regardless of status
            total_inventory_by_grade[grade] += 1
            
            # Record sales by month
            if stock == 'SOLD' and sale_date:
                sale_month = sale_date.strftime('%Y-%m')
                sold_by_month_and_grade[sale_month][grade] += 1
                all_months.add(sale_month)

        sorted_months = sorted(all_months)
        
        # Calculate stock at the end of each month
        stock_by_month = defaultdict(lambda: defaultdict(int))
        cumulative_sold_by_grade = defaultdict(int)
        
        for month in sorted_months:
            for grade in grades_seen:
                # Add sales for this month to the running total
                sold_this_month = sold_by_month_and_grade[month].get(grade, 0)
                cumulative_sold_by_grade[grade] += sold_this_month
                
                # Stock at end of month = total inventory - cumulative sold
                stock_by_month[month][grade] = total_inventory_by_grade[grade] - cumulative_sold_by_grade[grade]

        # --- PDF Generation ---
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        y = height - 50

        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, y, "Sales & Stock Report by Grade")
        y -= 30

        p.setFont("Helvetica", 12)
        for month in sorted_months:
            p.drawString(50, y, f"Month: {month}")
            y -= 20

            # SOLD section
            p.drawString(70, y, "SOLD:")
            y -= 15
            sold_total = 0
            for grade in sorted(grades_seen):
                qty = sold_by_month_and_grade[month].get(grade, 0)
                sold_total += qty
                p.drawString(90, y, f"Grade {grade}: {qty}")
                y -= 15
            p.drawString(90, y, f"total: {sold_total}")
            y -= 25

            # In Stock section
            p.drawString(70, y, "In Stock (End of Month):")
            y -= 15
            stock_total = 0
            for grade in sorted(grades_seen):
                qty = stock_by_month[month].get(grade, 0)
                stock_total += qty
                p.drawString(90, y, f"Grade {grade}: {qty}")
                y -= 15
            p.drawString(90, y, f"total: {stock_total}")
            y -= 30

            if y < 100:
                p.showPage()
                y = height - 50
                p.setFont("Helvetica", 12)

        p.save()
        buffer.seek(0)

        response = make_response(buffer.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'attachment; filename=sales_stock_report.pdf'
        return response
    except Exception as e:
        return f"Error: {str(e)}"

@laptop_bp.route('/laptop_ARP', methods=['GET'])
def laptop_ARP():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT Stock, OrderNumber FROM RMA_laptop_sheet")
        rows = cursor.fetchall()
        conn.close()

        sold_rows = [row for row in rows if row[0] == "SOLD"]
        sold_count = len(sold_rows)

        # Create Excel in memory
        wb = Workbook()
        ws = wb.active
        ws.title = "ARP Stock Report"

        # Header
        ws["A1"] = "ARP Stock Report"
        ws["A3"] = "SOLD:"
        ws["B3"] = sold_count
        ws["A5"] = "Order Numbers"

        # Add Order Numbers
        for idx, (_, order_number) in enumerate(sold_rows, start=6):
            ws[f"A{idx}"] = order_number

        # Save to memory
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name="arp_report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return f"Error: {str(e)}"


    
@laptop_bp.route('/laptop_TRP', methods=['GET'])
def laptop_TRP():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT InputDate, TechDoneDate FROM RMA_laptop_sheet")
        rows = cursor.fetchall()
        conn.close()

        input_date_group = defaultdict(int) #all laptop
        quickcheck_date_group = defaultdict(int) #quick check

        for row in rows:
            input_date, quickcheck_date = row

            if input_date:
                input_month = input_date.strftime('%Y-%m')
                input_date_group[input_month] += 1
                
            if quickcheck_date:
                tech_month = quickcheck_date.strftime('%Y-%m')
                quickcheck_date_group[tech_month] += 1
        
        # all laptop = included sealed and not sealed
        # quick check = only sealed laptop
        final_result = {
            key: input_date_group[key] - quickcheck_date_group.get(key, 0)
            for key in input_date_group
        }
        
        data = {
            'Full_qty': dict(final_result),
            'Quick_check_qty': dict(quickcheck_date_group)
        }
        
        # Create PDF in memory
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter

        p.setFont("Helvetica", 14)
        p.drawString(100, height - 50, "TRP Monthly Report")

        y = height - 100
        p.setFont("Helvetica", 12)

        # Draw Full_qty section
        p.drawString(50, y, "Full_qty:")
        y -= 20
        for month, qty in data["Full_qty"].items():
            p.drawString(70, y, f"{month}: {qty}")
            y -= 20

        y -= 10
        p.drawString(50, y, "Quick_check_qty:")
        y -= 20
        for month, qty in data["Quick_check_qty"].items():
            p.drawString(70, y, f"{month}: {qty}")
            y -= 20

        p.showPage()
        p.save()

        buffer.seek(0)

        # Send as response
        response = make_response(buffer.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'attachment; filename=trp_report.pdf'
        return response
    except Exception as e:
        return f"Error: {str(e)}"