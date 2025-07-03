# app/routes/report.py
from flask import Blueprint, request, jsonify, make_response, send_file
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from io import BytesIO
from collections import defaultdict
from app.models import get_db_connection
from openpyxl import Workbook
from openpyxl.styles import Font

report_bp = Blueprint('report', __name__)

@report_bp.route('/laptop_TRP', methods=['GET'])
def laptop_TRP():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT InputDate, TechDoneDate FROM RMA_laptop_sheet")
        rows = cursor.fetchall()
        
        cursor.execute("SELECT InputDate, Condition, InspectionRequest, Category FROM RMA_non_laptop_sheet")
        rows_nonpc = cursor.fetchall()

        conn.close()

        input_date_group = defaultdict(int)
        quickcheck_date_group = defaultdict(int)

        for input_date, quickcheck_date in rows:
            if input_date:
                input_month = input_date.strftime('%Y-%m')
                input_date_group[input_month] += 1
            if quickcheck_date:
                tech_month = quickcheck_date.strftime('%Y-%m')
                quickcheck_date_group[tech_month] += 1

        final_result = {
            key: input_date_group[key] - quickcheck_date_group.get(key, 0)
            for key in input_date_group
        }

        data = {
            'Detail_Check_Qty': dict(final_result),
            'Brand_new_Qty': dict(quickcheck_date_group)
        }
        
        
        brand_new_group = defaultdict(int)
        quick_check_group = defaultdict(int)
        detail_check_group = defaultdict(int)

        for input_date, condition, inspection_request, category in rows_nonpc:
            if input_date:
                input_month = input_date.strftime('%Y-%m')
                if condition == 'W':
                    brand_new_group[input_month] += 1
                elif inspection_request == 'A' or category == 'Gaming Console':
                    detail_check_group[input_month] += 1
                else:
                    quick_check_group[input_month] += 1

        data_nonPC = {
            'Brand_new_Qty': dict(brand_new_group),
            'Quick_Check_Qty': dict(quick_check_group),
            'Detail_Check_Qty': dict(detail_check_group)
        }

        export_format = request.args.get('format', 'pdf')

        if export_format == 'excel':
            # Excel export using openpyxl
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "TRP Report"

            ws.append([
                "Month",
                "PC: Detail_Check_Qty",
                "PC: Brand_new_Qty",
                "NonPC: Brand_new_Qty",
                "NonPC: Quick_Check_Qty",
                "NonPC: Detail_Check_Qty"
            ])
            all_months = sorted(
                set(data["Detail_Check_Qty"].keys())
                | set(data["Brand_new_Qty"].keys())
                | set(data_nonPC["Brand_new_Qty"].keys())
                | set(data_nonPC["Quick_Check_Qty"].keys())
                | set(data_nonPC["Detail_Check_Qty"].keys())
            )
            for month in all_months:
                ws.append([
                    month,
                    data["Detail_Check_Qty"].get(month, 0),
                    data["Brand_new_Qty"].get(month, 0),
                    data_nonPC["Brand_new_Qty"].get(month, 0),
                    data_nonPC["Quick_Check_Qty"].get(month, 0),
                    data_nonPC["Detail_Check_Qty"].get(month, 0),
                ])
            wb.save(output)
            output.seek(0)

            return send_file(output, as_attachment=True, download_name="trp_report.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        else:
            # PDF export
            buffer = BytesIO()
            p = canvas.Canvas(buffer, pagesize=letter)
            width, height = letter

            p.setFont("Helvetica", 14)
            p.drawString(100, height - 50, "TRP Monthly Report")
            y = height - 100
            p.setFont("Helvetica", 12)

            # PC section
            p.drawString(50, y, "PC: Detail Check Qty:")
            y -= 20
            for month, qty in data["Detail_Check_Qty"].items():
                p.drawString(70, y, f"{month}: {qty}")
                y -= 20

            y -= 10
            p.drawString(50, y, "PC: Brand New Qty:")
            y -= 20
            for month, qty in data["Brand_new_Qty"].items():
                p.drawString(70, y, f"{month}: {qty}")
                y -= 20

            # Non-PC section
            y -= 20
            p.drawString(50, y, "Non-PC: Brand New Qty:")
            y -= 20
            for month, qty in data_nonPC["Brand_new_Qty"].items():
                p.drawString(70, y, f"{month}: {qty}")
                y -= 20

            y -= 10
            p.drawString(50, y, "Non-PC: Quick Check Qty:")
            y -= 20
            for month, qty in data_nonPC["Quick_Check_Qty"].items():
                p.drawString(70, y, f"{month}: {qty}")
                y -= 20

            y -= 10
            p.drawString(50, y, "Non-PC: Detail Check Qty:")
            y -= 20
            for month, qty in data_nonPC["Detail_Check_Qty"].items():
                p.drawString(70, y, f"{month}: {qty}")
                y -= 20

            p.showPage()
            p.save()

            buffer.seek(0)
            response = make_response(buffer.read())
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = 'attachment; filename=trp_report.pdf'
            return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"Error: {str(e)}", 500

@report_bp.route('/laptop_ARP', methods=['GET'])
def laptop_ARP():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT Stock, OrderNumber, SaleDate FROM RMA_laptop_sheet")
        rows = cursor.fetchall()
        conn.close()

        sold_rows = [row for row in rows if row[0] == "SOLD"]
        sold_count = len(sold_rows)

        # Group sold order numbers by sale month
        grouped = defaultdict(list)
        for stock, order_number, sale_date in sold_rows:
            if sale_date:
                month = sale_date.strftime('%Y-%m')
            else:
                month = "No Sale Date"
            grouped[month].append(order_number)

        export_format = request.args.get('format', 'excel')

        if export_format == 'pdf':
            buffer = BytesIO()
            p = canvas.Canvas(buffer, pagesize=letter)
            width, height = letter

            p.setFont("Helvetica", 16)
            p.drawString(100, height - 50, "ARP Stock Report")

            p.setFont("Helvetica", 12)
            y = height - 100
            p.drawString(50, y, f"Total SOLD: {sold_count}")
            y -= 30

            for month, orders in grouped.items():
                p.drawString(50, y, f"{month}: Order Numbers")
                y -= 20
                for order in orders:
                    p.drawString(70, y, str(order))
                    y -= 20
                    if y < 50:
                        p.showPage()
                        y = height - 50
                y -= 10

            p.showPage()
            p.save()
            buffer.seek(0)

            return send_file(
                buffer,
                as_attachment=True,
                download_name="arp_report.pdf",
                mimetype="application/pdf"
            )

        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "ARP Stock Report"

            ws["A1"] = "ARP Stock Report"
            ws["A3"] = "Total SOLD:"
            ws["B3"] = sold_count

            row_idx = 5
            for month, orders in grouped.items():
                ws[f"A{row_idx}"] = f"{month}: Order Numbers"
                row_idx += 1
                for order in orders:
                    ws[f"A{row_idx}"] = order
                    row_idx += 1
                row_idx += 1  # add spacing between months

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return send_file(
                buffer,
                as_attachment=True,
                download_name="arp_report.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        return f"Error: {str(e)}", 500

@report_bp.route('/laptop_SRP', methods=['GET'])
def laptop_SRP():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT SaleDate, Stock, Condition FROM RMA_laptop_sheet")
        rows = cursor.fetchall()
        conn.close()

        # Grading Labels
        grade_labels = {
            'N': 'back to new',
            'A': 'Grade A',
            'B': 'Grade B',
            'C': 'Grade C',
            'F': 'Grade F'
        }
        valid_grades = set(grade_labels.keys())
        grades_seen = set()

        # Result containers
        sold_by_month_grade = defaultdict(lambda: defaultdict(int))  # monthly SOLD
        in_stock_by_grade = defaultdict(int)  # not SOLD
        total_inventory_by_grade = defaultdict(int)  # total per grade
        all_months = set()

        for sale_date, stock, grade in rows:
            grade = (grade or '').strip().upper()
            if grade not in valid_grades:
                continue

            grades_seen.add(grade)

            # Count total inventory per grade
            total_inventory_by_grade[grade] += 1

            # SOLD
            if stock == "SOLD":
                if sale_date:
                    month = sale_date.strftime('%Y-%m')
                    sold_by_month_grade[month][grade] += 1
                    all_months.add(month)
                else:
                    sold_by_month_grade["No date"][grade] += 1
            else:
                # Not sold — still in stock
                in_stock_by_grade[grade] += 1

        # Ensure "No date" appears last in reports
        sorted_months = sorted(m for m in all_months)
        if "No date" in sold_by_month_grade:
            sorted_months.append("No date")

        # Total sold count
        total_sold_qty = sum(
            sum(grade_count.values()) for month, grade_count in sold_by_month_grade.items()
        )

        export_format = request.args.get('format', 'pdf')

        if export_format == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = "Sales & Stock Report"

            row = 1
            ws["A1"] = "Sales & Stock Report by Grade"
            ws["A2"] = f"Total SOLD Quantity: {total_sold_qty}"
            ws["A1"].font = Font(bold=True)
            ws["A2"].font = Font(bold=True)
            row = 4

            for month in sorted_months:
                ws[f"A{row}"] = f"{month}:"
                ws[f"A{row}"].font = Font(bold=True)
                row += 1
                for grade in sorted(grades_seen):
                    label = grade_labels[grade]
                    qty = sold_by_month_grade[month].get(grade, 0)
                    ws[f"A{row}"] = label
                    ws[f"B{row}"] = qty
                    row += 1
                row += 1

            row += 1
            ws[f"A{row}"] = "Sum of IN STOCK laptops:"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
            for grade in sorted(grades_seen):
                label = grade_labels[grade]
                qty = in_stock_by_grade.get(grade, 0)
                ws[f"A{row}"] = label
                ws[f"B{row}"] = qty
                row += 1

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return send_file(
                buffer,
                as_attachment=True,
                download_name="sales_stock_report.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        else:
            # (PDF code unchanged)
            buffer = BytesIO()
            p = canvas.Canvas(buffer, pagesize=letter)
            width, height = letter
            y = height - 50

            p.setFont("Helvetica-Bold", 16)
            p.drawString(50, y, "Sales & Stock Report by Grade")
            y -= 30

            p.setFont("Helvetica", 12)
            p.drawString(50, y, f"Total SOLD Quantity: {total_sold_qty}")
            y -= 30

            for month in sorted_months:
                p.setFont("Helvetica-Bold", 12)
                p.drawString(50, y, f"{month}:")
                y -= 20
                p.setFont("Helvetica", 12)

                for grade in sorted(grades_seen):
                    qty = sold_by_month_grade[month].get(grade, 0)
                    label = grade_labels[grade]
                    p.drawString(70, y, f"{label}: {qty}")
                    y -= 15

                y -= 10
                if y < 100:
                    p.showPage()
                    y = height - 50

            y -= 10
            if y < 150:
                p.showPage()
                y = height - 50

            p.setFont("Helvetica-Bold", 12)
            p.drawString(50, y, "Sum of IN STOCK laptops:")
            y -= 20
            p.setFont("Helvetica", 12)
            for grade in sorted(grades_seen):
                label = grade_labels[grade]
                qty = in_stock_by_grade.get(grade, 0)
                p.drawString(70, y, f"{label}: {qty}")
                y -= 15

            p.save()
            buffer.seek(0)

            response = make_response(buffer.read())
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = 'attachment; filename=sales_stock_report.pdf'
            return response

    except Exception as e:
        return f"Error: {str(e)}", 500
